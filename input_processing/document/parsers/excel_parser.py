# risk_rag_system/input_processing/document/parsers/excel_parser.py

from typing import Dict, Any, List, Optional, Tuple
from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from loguru import logger
from pydantic import BaseModel
from PIL import Image
import io
import pytesseract
from datetime import datetime

class ExcelParserConfig(BaseModel):
    """Configuration for Excel parsing"""
    extract_images: bool = True
    perform_ocr: bool = True
    extract_formulas: bool = True
    extract_charts: bool = True
    extract_hidden_sheets: bool = False
    max_rows: Optional[int] = None
    max_cols: Optional[int] = None
    handle_merged_cells: bool = True
    date_format: str = "%Y-%m-%d"
    numeric_precision: int = 6

class ExcelParser:
    """Parser for Excel documents"""
    
    def __init__(self, config: Optional[ExcelParserConfig] = None):
        self.config = config or ExcelParserConfig()
        logger.info("Initialized ExcelParser")

    async def parse(
        self,
        file_path: Path,
        metadata: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """Parse Excel document"""
        try:
            logger.info(f"Parsing Excel: {file_path}")
            
            # Load workbook for both data and metadata extraction
            workbook = load_workbook(
                filename=file_path,
                data_only=True,  # Get values instead of formulas
                read_only=not self.config.extract_images  # Read-only if not extracting images
            )
            
            # Load workbook with formulas if needed
            if self.config.extract_formulas:
                formula_wb = load_workbook(filename=file_path, data_only=False)
            
            # Process each worksheet
            content = []
            tables = []
            images = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Skip hidden sheets if configured
                if not self.config.extract_hidden_sheets and sheet.sheet_state == 'hidden':
                    continue
                
                # Process worksheet data
                sheet_data = self._process_worksheet(
                    sheet,
                    formula_wb[sheet_name] if self.config.extract_formulas else None
                )
                
                if sheet_data:
                    tables.append({
                        "name": sheet_name,
                        "data": sheet_data["data"],
                        "formulas": sheet_data.get("formulas", {}),
                        "metadata": {
                            "num_rows": len(sheet_data["data"]),
                            "num_cols": len(sheet_data["data"][0]) if sheet_data["data"] else 0,
                            "sheet_name": sheet_name
                        }
                    })
                    
                    # Add text representation to content
                    content.append(f"Sheet: {sheet_name}")
                    content.append(self._table_to_text(sheet_data["data"]))
                
                # Extract images if enabled
                if self.config.extract_images:
                    sheet_images = self._extract_images(sheet)
                    images.extend(sheet_images)
            
            # Process images with OCR if enabled
            if self.config.perform_ocr and images:
                ocr_text = await self._process_ocr(images)
                if ocr_text:
                    content.append(ocr_text)
            
            # Get workbook properties
            wb_props = self._get_workbook_properties(workbook)
            
            # Prepare result
            result = {
                "content": "\n\n".join(content),
                "metadata": {
                    "source": str(file_path),
                    "type": "excel",
                    **wb_props,
                    **(metadata or {})
                },
                "tables": tables,
                "images": images
            }
            
            # Clean up
            workbook.close()
            if self.config.extract_formulas:
                formula_wb.close()
            
            return result
            
        except Exception as e:
            logger.error(f"Error parsing Excel {file_path}: {e}")
            raise

    def _process_worksheet(
        self,
        sheet,
        formula_sheet: Optional[Any] = None
    ) -> Dict[str, Any]:
        """Process worksheet data and formulas"""
        try:
            data = []
            formulas = {}
            
            # Get dimensions
            max_row = min(sheet.max_row, self.config.max_rows) if self.config.max_rows else sheet.max_row
            max_col = min(sheet.max_column, self.config.max_cols) if self.config.max_cols else sheet.max_column
            
            # Handle merged cells if configured
            merged_ranges = {}
            if self.config.handle_merged_cells:
                for merged_range in sheet.merged_cells.ranges:
                    merged_ranges[merged_range.coord] = merged_range
            
            # Process cells
            for row in range(1, max_row + 1):
                row_data = []
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell_value = self._process_cell_value(cell)
                    
                    # Handle merged cells
                    if self.config.handle_merged_cells:
                        for merged_range in merged_ranges.values():
                            if cell.coordinate in merged_range:
                                if cell.coordinate == merged_range.start_cell.coordinate:
                                    cell_value = self._process_cell_value(
                                        sheet.cell(
                                            row=merged_range.min_row,
                                            column=merged_range.min_col
                                        )
                                    )
                                else:
                                    cell_value = None
                                break
                    
                    row_data.append(cell_value)
                    
                    # Extract formulas if enabled
                    if self.config.extract_formulas and formula_sheet:
                        formula_cell = formula_sheet.cell(row=row, column=col)
                        if formula_cell.data_type == 'f':
                            formulas[cell.coordinate] = formula_cell.value
                
                if any(x is not None for x in row_data):
                    data.append(row_data)
            
            return {
                "data": data,
                "formulas": formulas if self.config.extract_formulas else None
            }
            
        except Exception as e:
            logger.error(f"Error processing worksheet: {e}")
            return {"data": [], "formulas": {}}

    def _process_cell_value(self, cell) -> Any:
        """Process cell value with proper formatting"""
        if cell.value is None:
            return None
            
        if isinstance(cell.value, datetime):
            return cell.value.strftime(self.config.date_format)
            
        if isinstance(cell.value, (int, float)):
            if np.isnan(cell.value):
                return None
            return round(cell.value, self.config.numeric_precision)
            
        return str(cell.value)

    def _extract_images(self, sheet) -> List[Dict[str, Any]]:
        """Extract images from worksheet"""
        try:
            images = []
            
            for image in sheet._images:
                try:
                    if isinstance(image, XLImage):
                        image_data = {
                            "content": image._data(),
                            "size": (image.width, image.height),
                            "format": image.format.lower() if hasattr(image, 'format') else 'unknown',
                            "sheet": sheet.title,
                            "index": len(images)
                        }
                        images.append(image_data)
                        
                except Exception as img_error:
                    logger.warning(f"Error processing worksheet image: {img_error}")
                    continue
                    
            return images
            
        except Exception as e:
            logger.error(f"Error extracting worksheet images: {e}")
            return []

    async def _process_ocr(self, images: List[Dict[str, Any]]) -> str:
        """Process OCR on extracted images"""
        try:
            ocr_results = []
            
            for img_data in images:
                image = Image.open(io.BytesIO(img_data["content"]))
                text = pytesseract.image_to_string(image)
                
                if text.strip():
                    ocr_results.append(f"Image OCR (Sheet: {img_data['sheet']}):")
                    ocr_results.append(text)
            
            return "\n\n".join(ocr_results)
            
        except Exception as e:
            logger.error(f"Error processing OCR: {e}")
            return ""

    def _table_to_text(self, data: List[List[Any]]) -> str:
        """Convert table data to text representation"""
        if not data:
            return ""
            
        # Convert all values to strings
        str_data = [[str(val) if val is not None else "" for val in row] for row in data]
        
        # Get column widths
        col_widths = []
        for col in range(len(str_data[0])):
            col_widths.append(max(len(str(row[col])) for row in str_data))
        
        # Build text representation
        lines = []
        
        # Header
        lines.append(" | ".join(
            str_data[0][i].ljust(col_widths[i])
            for i in range(len(col_widths))
        ))
        
        # Separator
        lines.append("-" * (sum(col_widths) + 3 * (len(col_widths) - 1)))
        
        # Data rows
        for row in str_data[1:]:
            lines.append(" | ".join(
                str(row[i]).ljust(col_widths[i])
                for i in range(len(col_widths))
            ))
        
        return "\n".join(lines)

    def _get_workbook_properties(self, workbook) -> Dict[str, Any]:
        """Extract workbook properties"""
        try:
            props = workbook.properties
            return {
                "title": props.title or "",
                "subject": props.subject or "",
                "creator": props.creator or "",
                "created": props.created.isoformat() if props.created else "",
                "modified": props.modified.isoformat() if props.modified else "",
                "last_modified_by": props.lastModifiedBy or "",
                "sheets": workbook.sheetnames,
                "num_sheets": len(workbook.sheetnames)
            }
        except Exception as e:
            logger.error(f"Error getting workbook properties: {e}")
            return {}